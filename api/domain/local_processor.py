"""Local document processor — runs OCR/extraction without S3 or Postgres.

Processes files directly from the workspace filesystem and updates SQLite.
"""

import asyncio
import json
import logging
from pathlib import Path

import aiosqlite

logger = logging.getLogger(__name__)

TEXT_TYPES = frozenset({"md", "txt", "csv", "html", "svg", "json", "xml", "yaml", "yml"})
PDF_TYPES = frozenset({"pdf"})
SPREADSHEET_TYPES = frozenset({"xlsx", "xls"})
IMAGE_TYPES = frozenset({"png", "jpg", "jpeg", "webp", "gif"})


async def process_document(db: aiosqlite.Connection, doc_id: str, workspace: Path) -> None:
    """Process a pending document: extract text, chunk, update index."""
    cursor = await db.execute(
        "SELECT id, filename, file_type, relative_path, status FROM documents WHERE id = ?",
        (doc_id,),
    )
    row = await cursor.fetchone()
    if not row:
        logger.warning("Document %s not found", doc_id[:8])
        return

    cols = [d[0] for d in cursor.description]
    doc = dict(zip(cols, row))

    if doc["status"] not in ("pending", "processing"):
        return

    file_type = doc["file_type"] or ""
    file_path = workspace / doc["relative_path"]

    if not file_path.is_file():
        await db.execute(
            "UPDATE documents SET status = 'failed', error_message = 'File not found', "
            "updated_at = datetime('now') WHERE id = ?",
            (doc_id,),
        )
        await db.commit()
        return

    await db.execute(
        "UPDATE documents SET status = 'processing', updated_at = datetime('now') WHERE id = ?",
        (doc_id,),
    )
    await db.commit()

    try:
        if file_type in PDF_TYPES:
            await _process_pdf(db, doc_id, file_path)
        elif file_type in SPREADSHEET_TYPES:
            await _process_spreadsheet(db, doc_id, file_path)
        elif file_type in IMAGE_TYPES:
            await _process_image(db, doc_id, file_path)
        elif file_type in ("html", "htm"):
            await _process_html(db, doc_id, file_path)
        else:
            # Already indexed as text or unsupported — mark ready
            await db.execute(
                "UPDATE documents SET status = 'ready', updated_at = datetime('now') WHERE id = ?",
                (doc_id,),
            )
            await db.commit()

        logger.info("Processed %s: %s", doc["filename"], file_type)

    except Exception as e:
        error_msg = str(e)[:500]
        await db.execute(
            "UPDATE documents SET status = 'failed', error_message = ?, "
            "updated_at = datetime('now') WHERE id = ?",
            (error_msg, doc_id),
        )
        await db.commit()
        logger.error("Failed to process %s: %s", doc["filename"], e)


async def _process_pdf(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract PDF text via pdf-oxide."""
    from services.ocr import OCRService

    page_contents = await asyncio.to_thread(
        OCRService._extract_pdf_oxide, str(file_path),
    )
    num_pages = len(page_contents)

    # Store pages
    await db.execute("DELETE FROM document_pages WHERE document_id = ?", (doc_id,))
    import uuid
    for page_num, content in page_contents:
        await db.execute(
            "INSERT INTO document_pages (id, document_id, page, content) VALUES (?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, page_num, content),
        )

    # Build full content + chunks
    full_content = "\n\n---\n\n".join(md for _, md in page_contents)

    from services.chunker import chunk_pages
    chunks = chunk_pages(page_contents)

    # Store chunks
    await db.execute("DELETE FROM document_chunks WHERE document_id = ?", (doc_id,))
    for c in chunks:
        await db.execute(
            "INSERT INTO document_chunks (id, document_id, chunk_index, content, page, "
            "start_char, token_count, header_breadcrumb) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, c.index, c.content, c.page,
             c.start_char, c.token_count, c.header_breadcrumb),
        )

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = ?, "
        "parser = 'pdf_oxide', updated_at = datetime('now') WHERE id = ?",
        (full_content, num_pages, doc_id),
    )
    await db.commit()


async def _process_spreadsheet(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract spreadsheet data via openpyxl."""
    import uuid
    from openpyxl import load_workbook

    wb = await asyncio.to_thread(load_workbook, str(file_path), read_only=True, data_only=True)

    await db.execute("DELETE FROM document_pages WHERE document_id = ?", (doc_id,))

    all_content = []
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(" | ".join(str(c) if c is not None else "" for c in row))
        content = "\n".join(rows)
        elements = json.dumps({"sheet_name": sheet_name})

        await db.execute(
            "INSERT INTO document_pages (id, document_id, page, content, elements) "
            "VALUES (?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, i, content, elements),
        )
        all_content.append(f"## {sheet_name}\n\n{content}")

    wb.close()
    full_content = "\n\n".join(all_content)

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = ?, "
        "parser = 'openpyxl', updated_at = datetime('now') WHERE id = ?",
        (full_content, len(wb.sheetnames), doc_id),
    )
    await db.commit()


async def _process_image(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Images are stored as-is — just mark ready."""
    await db.execute(
        "UPDATE documents SET status = 'ready', page_count = 1, "
        "parser = 'native', updated_at = datetime('now') WHERE id = ?",
        (doc_id,),
    )
    await db.commit()


async def _process_html(db: aiosqlite.Connection, doc_id: str, file_path: Path) -> None:
    """Extract HTML content via webmd parser."""
    import uuid
    raw_html = file_path.read_text(encoding="utf-8", errors="replace")

    try:
        from html_parser import Parser
        parser = Parser(raw_html, content_only=True)
        result = parser.parse()
        content = result.content
    except Exception:
        # Fallback: store raw HTML as content
        content = raw_html

    from services.chunker import chunk_text
    chunks = chunk_text(content)

    await db.execute("DELETE FROM document_chunks WHERE document_id = ?", (doc_id,))
    for c in chunks:
        await db.execute(
            "INSERT INTO document_chunks (id, document_id, chunk_index, content, page, "
            "start_char, token_count, header_breadcrumb) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (str(uuid.uuid4()), doc_id, c.index, c.content, c.page,
             c.start_char, c.token_count, c.header_breadcrumb),
        )

    await db.execute(
        "UPDATE documents SET status = 'ready', content = ?, page_count = 1, "
        "parser = 'webmd', updated_at = datetime('now') WHERE id = ?",
        (content, doc_id),
    )
    await db.commit()
